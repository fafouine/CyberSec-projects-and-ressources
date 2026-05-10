"""
Excel metadata handler for Excel files.

This module provides the ExcelHandler class which implements the MetadataHandler
interface for Excel files (.xlsx, .xlsm, .xltx, .xltm). Uses openpyxl for
reading and writing Excel workbook properties.

Note:
    Does not support password-protected/encrypted workbooks.
"""

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.services.metadata_handler import MetadataHandler
from src.utils.exceptions import (
    MetadataNotFoundError,
    MetadataReadingError,
    UnsupportedFormatError,
)

# Supported Excel formats
FORMAT_MAP = {
    "xlsx": "xlsx",
    "xlsm": "xlsm",
    "xltx": "xltx",
    "xltm": "xltm",
}

# Properties to preserve (not deleted during wipe)
PRESERVED_PROPERTIES = {"created", "modified", "language"}


class ExcelHandler(MetadataHandler):
    """
    Excel metadata handler for Excel files.

    Handles extraction and removal of document properties from Excel workbooks
    including author, title, subject, keywords, and other core properties.

    Attributes:
        keys_to_delete: List of property names to be wiped.
    """
    def __init__(self, filepath: str):
        """
        Initialize the Excel handler.

        Args:
            filepath: Path to the Excel file to process.
        """
        super().__init__(filepath)
        self.keys_to_delete: list[str] = []

    def _detect_format(self) -> str:
        """
        Detect Excel format from file extension.

        Returns:
            Normalized format string ('xlsx', 'xlsm', 'xltx', or 'xltm').

        Raises:
            UnsupportedFormatError: If file extension is not a supported Excel format.
        """
        ext = Path(self.filepath).suffix.lower()
        normalised = FORMAT_MAP.get(ext[1 :])  # Remove leading dot
        if normalised is None:
            raise UnsupportedFormatError(f"Unsupported format: {ext}")

        return normalised

    def read(self) -> dict[str, Any]:
        """
        Extract metadata properties from the Excel workbook.

        Reads all document properties from the workbook and identifies
        which properties should be wiped (excludes created, modified, language).

        Returns:
            Dictionary of property names to their values.

        Raises:
            MetadataReadingError: If the workbook is password-protected.
            MetadataNotFoundError: If no properties are found.
        """
        self.metadata.clear()
        self.keys_to_delete.clear()
        wb = load_workbook(Path(self.filepath))
        try:
            if wb.security.workbookPassword is not None:
                raise MetadataReadingError("File is encrypted.")

            if wb.properties is None:
                raise MetadataNotFoundError("No metadata found in the file.")

            for attr, value in vars(wb.properties).items():
                self.metadata[attr] = value
                if attr not in PRESERVED_PROPERTIES:
                    self.keys_to_delete.append(attr)

            return self.metadata
        finally:
            wb.close()

    def wipe(self) -> None:
        """
        Remove metadata properties from the Excel workbook.

        Clears all properties identified during read() except for
        preserved properties (created, modified, language).

        Raises:
            MetadataNotFoundError: If no properties are found.
        """
        self.processed_metadata.clear()
        wb = load_workbook(Path(self.filepath))
        try:
            if wb.properties is None:
                raise MetadataNotFoundError("No metadata found in the file.")

            # Clear each property marked for deletion
            for attr in self.keys_to_delete:
                if hasattr(wb.properties, attr):
                    setattr(wb.properties, attr, None)

            self.processed_metadata = wb.properties
        finally:
            wb.close()

    def save(self, output_path: str | None) -> None:
        """
        Save the workbook with cleaned metadata to the output path.

        Creates a copy of the original file and applies the wiped
        metadata properties to it.

        Args:
            output_path: Path where the cleaned file should be saved.

        Raises:
            ValueError: If output_path is None or empty.
        """
        if not output_path:
            raise ValueError("output_path is required")

        destination_file_path = Path(output_path)
        shutil.copy2(self.filepath, destination_file_path)

        # Use keep_vba=True for macro-enabled workbooks
        detected_format = self._detect_format()
        if detected_format == "xlsm":
            wb = load_workbook(destination_file_path, keep_vba = True)
        else:
            wb = load_workbook(destination_file_path)

        try:
            # Apply wiped properties
            for attr, value in vars(self.processed_metadata).items():
                setattr(wb.properties, attr, value)

            wb.save(destination_file_path)
        finally:
            wb.close()
